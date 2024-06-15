import pandas as pd
from openpyxl import load_workbook

def add_course(subject, class_name, assignment, weightage):
    # print("function called")
    wb = load_workbook("Hackathon Berlin.xlsx", read_only=True)
    pd_writer = pd.ExcelWriter("Hackathon Berlin.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay")
    new = pd.DataFrame({
        "subject" : [subject],
        "class_name": [class_name],
    })     
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    df = pd.read_excel(xlsx, 'Course')
    # print(df.empty)
    if "Course" not in wb.sheetnames or df.empty:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Course", index=False)
    else:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Course", startrow=writer.sheets["Course"].max_row, index=False, header=False)
    
    new_assignment = pd.DataFrame({
        "class_name": [class_name],
        "assignment": [assignment],
        "weightage": [weightage]
    })
    pd_writer = pd.ExcelWriter("Hackathon Berlin.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay")
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    df1 = pd.read_excel(xlsx, 'Assignment Weightage')
    if "Assignment Weightage" not in wb.sheetnames or df1.empty:
        with pd_writer as writer:
            new_assignment.to_excel(writer, sheet_name="Assignment Weightage", index=False)
    else:
        with pd_writer as writer:
            new_assignment.to_excel(writer, sheet_name="Assignment Weightage", startrow=writer.sheets["Assignment Weightage"].max_row, index=False, header=False)

    return True


def read_sheets():
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    a_df = pd.read_excel(xlsx, 'Assignments')
    s_df = pd.read_excel(xlsx, 'Subject')
    c_df = pd.read_excel(xlsx, 'Class')
    result = {}
    a_result = {}
    s_result = {}
    c_result = {}
    if not a_df.empty:
        a_result = a_df["Assignments"].to_dict()
    if not s_df.empty:
        s_result = s_df["Subject"].to_dict()
    if not c_df.empty:
        c_result = c_df["Class"].to_dict()
    result = {
        "Assignment" : a_result,
        "Subject" : s_result,
        "Class": c_result
    }
    # print(result)
    # subjects = result["subject"]
    # for key,values in subjects.items():
    #     print(values)
    return result

def read_courses():
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    df = pd.read_excel(xlsx, 'Course')
    result = {}
    if not df.empty:
        result = df.to_dict()
    # print(result['subject'])
    # subjects = result["subject"]
    # for key,values in subjects.items():
    #     print(values)
    return result

def add_subject(subject):
    pd_writer = pd.ExcelWriter("Hackathon Berlin.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay")
    new = pd.DataFrame({
        "Subject" : [subject],
    })     
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    df = pd.read_excel(xlsx, 'Subject')
    print(df.empty)
    if df.empty:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Subject", index=False)
    else:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Subject", startrow=writer.sheets["Subject"].max_row, index=False, header=False)
         
    return True

def add_class(class_):
    pd_writer = pd.ExcelWriter("Hackathon Berlin.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay")
    new = pd.DataFrame({
        "Class" : [class_],
    })     
    xlsx = pd.ExcelFile('Hackathon Berlin.xlsx')
    df = pd.read_excel(xlsx, 'Class')
    print(df.empty)
    if df.empty:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Class", index=False)
    else:
        with pd_writer as writer:
            new.to_excel(writer, sheet_name="Class", startrow=writer.sheets["Class"].max_row, index=False, header=False)
         
    return True


if __name__== "__main__":
    read_sheets()